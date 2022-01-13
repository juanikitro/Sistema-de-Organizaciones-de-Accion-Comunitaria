# Python
from datetime import date 

# Django
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.db.transaction import atomic
from django.views.generic import TemplateView
from django.db import IntegrityError
from django.conf import settings
from django.core.mail import send_mail

# Models
from django.contrib.auth.models import User
from users.models import Profile
from history.models import Item

# Open Py XL (Para el excel)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side, PatternFill
from django.http.response import HttpResponse

def login_view(request):
    ''' Login
    Se autentifica el usuario con username(cuit) y password con authenticate de Django
    Si el usuario no existe, imprime error '''

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password', False);
        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            return redirect('home')
        else:
            return render(request, 'users/login.html', {'error': 'Error en el cuit o la contraseña.'})

    return render(request, 'users/login.html')

@login_required
@atomic
def signup_view(request):
    '''Signup
    Values = void dict para que los inputs empiecen en 0 y se guarden para hacer cargas similares
    Se verifica que el username(cuit) y email no esten en uso
    Se guarda el usuario y el perfil.
    user similar
    Se usa atomic por si hay error en uno de los dos save() '''

    user_id = request.user.id
    profile_level = Profile.objects.get(user_id = user_id).level

    values = {}
    if request.method == 'POST':
        username = request.POST['username']        
        password = request.POST['password']
        email = request.POST['email']


        if Profile.objects.filter(username=username).first(): 
            return render(request, 'users/signup.html', {'error': 'El cuit ya pertenece a una cuenta'})
        if Profile.objects.filter(email=email).first(): 
            return render(request, 'users/signup.html', {'error': 'El email ya pertenece a una cuenta'})

        user = User.objects.create_user(username, email, password)
        user.username = username
        user.first_name = request.POST['first_name']
        user.last_name = request.POST['last_name']
        user.save()

        profile = Profile()
        profile.user = user
        profile.email = email
        profile.username = username
        profile.first_name = request.POST['first_name']
        profile.last_name = request.POST['last_name']
        profile.level = request.POST['level']
        profile.mobile = request.POST['mobile']
        profile.save()

        values={
            'name': request.POST['first_name'],
            'lastname': request.POST['last_name'],
            'cuit': username,
            'email': email,
            'mobile': request.POST['mobile'],
        }
        
        history_item = Item()
        history_item.action = f'Creacion de usuario: {profile.username}'
        history_item.by = f'{Profile.objects.get(user_id = user_id).first_name} {Profile.objects.get(user_id = user_id).last_name}'
        history_item.save()

        return render(request, 'users/signup.html', {'alert': 'Usuario creado con exito', 'level': profile_level})

    return render(request, 'users/signup.html', {'values': values, 'level': profile_level})

@login_required
def logout_view(request):
    '''Logout de Django... '''

    logout(request)
    return redirect('login')

@login_required
def users_view(request):
    '''Busqueda de usuarios
    Se crea la variable global profile(que sera util en otras vistas)
    Values = void dict para que los inputs empiecen en 0 y se guarden para hacer busquedas similares
    Prepara todos los datos para mostrar en la tabla 
    Se filtran los usuarios y se envian al template '''

    user_id = request.user.id
    profile_level = Profile.objects.get(user_id = user_id).level

    global profile
    profile = Profile.objects.all()
    values = {}
    if request.method == 'POST':
        name = request.POST.get('name', False)
        lastname = request.POST.get('lastname', False)
        cuit = request.POST.get('cuit', False)
        email = request.POST.get('email', False)
        mobile = request.POST.get('mobile', False)
        level = request.POST.get('level', False)

        if level == '1':
            level = 'comunal'
        elif level == '2':
            level = 'central'
        elif level == '3':
            level = 'presi'
        elif level == '0':
            level = ''

        values={
            'name': name,
            'lastname': lastname,
            'cuit': cuit,
            'email': email,
            'mobile': mobile,
        }
        profile = Profile.objects.filter(first_name__contains=name, last_name__contains=lastname, username__contains=cuit, email__contains=email, mobile__startswith=mobile, level__startswith=level)

    return render(request, 'users/users.html', {'profile': profile, 'values': values, 'level': profile_level})

class Excel_report(TemplateView):
    '''Export de usuarios
    Se crea el excel con los datos de los usuarios
    Se utilizan los usuarios filtrados por la variable global profile '''

    def get(self, *args, **kwargs):
        today = date.today()

        wb = Workbook()
        ws = wb.active
        ws['A1'] = f'Reporte de usuarios del dia: {today}'
        ws['A1'].alignment = Alignment(horizontal = 'center')
        ws['A1'].border = Border(left = Side(border_style = 'thin'), right = Side(border_style = 'thin'), bottom = Side(border_style = 'thin'), top = Side(border_style = 'thin'))
        ws['A1'].font = Font(name = 'Arial', size = 12)
        ws['A1'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')

        ws.merge_cells('A1:G1')

        ws['A3'] = 'id de usuario'
        ws['A3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')
        ws['B3'] = 'Cuit'
        ws['B3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')
        ws['C3'] = 'Nombre'
        ws['C3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')
        ws['D3'] = 'Apellido'
        ws['D3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')
        ws['E3'] = 'Email'
        ws['E3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')
        ws['F3'] = 'Contacto'
        ws['F3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')
        ws['G3'] = 'Nivel'
        ws['G3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 20

        cont = 4
        for u in profile:
            ws.cell(row = cont, column = 1).value = u.id
            ws.cell(row = cont, column = 2).value = u.username
            ws.cell(row = cont, column = 3).value = u.first_name
            ws.cell(row = cont, column = 4).value = u.last_name
            ws.cell(row = cont, column = 5).value = u.email
            ws.cell(row = cont, column = 6).value = u.mobile
            ws.cell(row = cont, column = 7).value = u.level
            cont += 1

        excel_name = f'Reporte de usuarios {today}.xlsx'
        response = HttpResponse(content_type = 'application/ms-excel')
        content = 'attachment; filename = {0}'.format(excel_name)
        response['Content-Disposition'] = content
        wb.save(response)
        return response

@login_required
def profile_view(request, pk):
    '''Vista del perfil 1:1 con usuarios
    Se busca que perfil coincide con la url y que usuario coincide con este perfil
    Se preparan todos los datos a mostrar en la tabla y se envian como context '''

    user_id = request.user.id
    profile_level = Profile.objects.get(user_id = user_id).level

    profile = Profile.objects.get(id=pk)
    user = User.objects.get(id=profile.user_id)

    id = profile.id
    username = user.username
    name = user.first_name + ' ' + user.last_name
    email = user.email
    mobile = profile.mobile
    created = profile.created
    modified = profile.modified
    level = ''
    if profile.level == 'admin':
        level = 'Administrador de ROAC'
    elif profile.level == 'presidente':
        level = 'Presidente'
    elif profile.level == 'central':
        level = 'Usuario de sede central'
    elif profile.level == 'comunal':
        level = 'Usuario de sede comunal'
    
    context = {
    'profile':profile,
    'id':id,
    'username':username,
    'name':name,
    'email':email,
    'mobile':mobile,
    'modified':modified,
    'created':created,
    'level':level,
    'level': profile_level
    }

    return render(request, 'users/profile.html', context)

@login_required
@atomic
def delete_profile_view(request, pk):
    '''Borrado de perfil
    Busca que usuario y perfil coincide con la url
    Elimina el usuario y perfil '''

    user_id = request.user.id
    selected_profile = Profile.objects.get(id=pk)
    selected_user = User.objects.get(id=selected_profile.user_id)

    history_item = Item()
    history_item.action = f'Eliminacion de usuario: {selected_profile.username}'
    history_item.by = f'{Profile.objects.get(user_id = user_id).first_name} {Profile.objects.get(user_id = user_id).last_name}'
    history_item.save()

    selected_profile.delete()
    selected_user.delete()

    return redirect('users')

@login_required
@atomic
def modify_profile_view(request, pk):
    '''Modificación de perfil
    Busca que usuario y perfil coincide con la url
    Guarda temporalmente la información vieja del perfil a modificar
    Extrae los datos del template para la modificación y previene el uso de un username(cuit) utilizado
    Sustituye los datos del usuario por los nuevamente colocados y guarda '''

    user_id = request.user.id
    profile_level = Profile.objects.get(user_id = user_id).level

    selected_profile = Profile.objects.get(id=pk)
    selected_user = User.objects.get(id=selected_profile.user_id)
    old_info = {
        'id': selected_profile.id,
        'username': selected_user.username,
        'email': selected_user.email,
        'name': selected_user.first_name,
        'lastname': selected_user.last_name,
        'mobile': selected_profile.mobile,
        'level': selected_profile.level,
    }

    if request.method == 'POST':
        selected_user.username = request.POST.get('username')     
        selected_user.first_name = request.POST['first_name']
        selected_user.last_name = request.POST['last_name']
        selected_user.email = request.POST['email']

        selected_profile.username = request.POST['username']    
        selected_profile.email = request.POST['email']
        selected_profile.first_name = request.POST['first_name']
        selected_profile.last_name = request.POST['last_name']
        selected_profile.level = request.POST['level']
        selected_profile.mobile = request.POST['mobile']

        try:
            history_item = Item()
            history_item.action = f'Modificacion de usuario: {selected_profile.username}'
            history_item.by = f'{Profile.objects.get(user_id = user_id).first_name} {Profile.objects.get(user_id = user_id).last_name}'
            history_item.save()

            selected_user.save()
            selected_profile.save()
        except IntegrityError:
            return render(request, 'users/modify_profile.html', {'old': old_info, 'error': 'El cuit o email ya esta en uso', 'level': profile_level})
            
        return redirect('profile', selected_profile.id)

    return render(request, 'users/modify_profile.html', {'old': old_info, 'level': profile_level})

@login_required
def reset_password_view(request, pk):
    '''Reseteo de password de Django
    Pide confirmación de password para evitar futuros problemas '''

    user_id = request.user.id
    profile_level = Profile.objects.get(user_id = user_id).level

    selected_profile = Profile.objects.get(id=pk)
    selected_user = User.objects.get(id=selected_profile.user_id)
    old_info = {
        'id': selected_profile.id,
        'username': selected_user.username,
        'name': selected_user.first_name,
        'lastname': selected_user.last_name,
    }

    if request.method == 'POST':
        new_password = request.POST.get('new_password')
        confirm_new_password = request.POST.get('confirm_new_password')

        if new_password == confirm_new_password:
            selected_user.set_password(new_password)

            history_item = Item()
            history_item.action = f'Blanqueo de clave de usuario: {selected_profile.username}'
            history_item.by = f'{Profile.objects.get(user_id = user_id).first_name} {Profile.objects.get(user_id = user_id).last_name}'
            history_item.save()

            selected_user.save()
            return redirect('profile', selected_profile.id)
        else:
            return render(request, 'users/reset_password.html', {'old': old_info, 'error': 'Las contraseñas no coinciden', 'level': profile_level})

    return render(request, 'users/reset_password.html', {'old': old_info, 'level': profile_level})

def send_reset_password_view(request):
    '''Reseteo de password para usuario no logeado
    Se pide cuit y se envia mail al usuario dueno del cuit
    Se le envia el mail para el link reset_password manual'''

    if request.method == 'POST':
        cuit = request.POST.get('username')
        profile_to_reset = Profile.objects.get(username=cuit)
        username_email = profile_to_reset.email
        link = f'http://127.0.0.1:8000/send_reset/sended/{profile_to_reset.id}/' #FIXME: Cambiar cuando existan los servers

        subject = f'Reseteo de contraseña para {profile_to_reset.first_name}'
        message = f'''Hola! Te contacto desde el Sistema de Organizaciones de Acción Comunitaria!

        Vimos que no podes entrar a tu cuenta en SOAC y necesitas un cambio de contraseña. Te envio el link para que puedas cambiarla! 

        {link}

        Saludos!
        '''
        email_from = settings.EMAIL_HOST_USER
        recipient_list = [username_email]

        if username_email:
            send_mail(subject, message, email_from, recipient_list)
            return render(request, 'users/send_reset_password.html', {'alert': 'Hemos enviado un link a tu correo electronico para que cambies la contraseña'})
        else:
            return render(request, 'users/send_reset_password.html', {'alert': 'No se ha podido enviar el mail, contacta con un usuario central / administrador para soluciónar'})

    return render(request, 'users/send_reset_password.html')

def reset_password_user_view(request, pk):
    '''Reseteo de contrasena manual para usuarios
    acceden mediante link en mail'''
    selected_profile = Profile.objects.get(id=pk)
    selected_user = User.objects.get(id=selected_profile.user_id)   
    if request.method == 'POST':
        new_password = request.POST.get('new_password')
        confirm_new_password = request.POST.get('confirm_new_password')    
        if new_password == confirm_new_password:
            selected_user.set_password(new_password)
            selected_user.save()
            return redirect('login')
        else:
            return render(request, 'users/reset_password_user.html', {'error': 'Las contraseñas no coinciden', 'user': selected_profile})  
    
    return render(request, 'users/reset_password_user.html', {'user': selected_profile})