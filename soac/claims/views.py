from datetime import date

#Django
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.views.generic import TemplateView

#Models
from claims.models import Claim
from users.models import Profile
from organizations.models import Org
from history.models import Item

# Open Py XL (Para el excel)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side, PatternFill
from django.http.response import HttpResponse

@login_required
def claims_view(request):
    user_id = request.user.id
    profile_level = Profile.objects.get(user_id = user_id).level

    global claims
    values = {}
    claims = Claim.objects.all().order_by('-created')

    if request.method == 'POST':
       by = request.POST.get('by', False)
       org_name = request.POST.get('org_name', False)
       observation = request.POST.get('observation', False)
       state = request.POST.get('state', False)

       values={
           'by': by,
           'org_name': org_name,
           'observation': observation,
           'state': state,
       }
       
       claims = Claim.objects.filter(state__contains=state, by__contains=by, org_name__contains=org_name, observation__contains=observation).order_by('-created')

    return render(request, 'claims/claims.html', {
        'level': profile_level,
        'claims': claims,
        'values':values
        })


class Excel_report(TemplateView):
    def get(self, *args, **kwargs):
        today = date.today()

        wb = Workbook()
        ws = wb.active
        ws['A1'] = f'Reporte de reclamos del dia: {today}'
        ws['A1'].alignment = Alignment(horizontal = 'center')
        ws['A1'].border = Border(left = Side(border_style = 'thin'), right = Side(border_style = 'thin'), bottom = Side(border_style = 'thin'), top = Side(border_style = 'thin'))
        ws['A1'].font = Font(name = 'Arial', size = 12)
        ws['A1'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')

        ws.merge_cells('A1:F1')

        ws['A3'] = 'id de reclamo'
        ws['A3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')
        ws['B3'] = 'Usuario'
        ws['B3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')
        ws['C3'] = 'Categoria'
        ws['C3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')
        ws['D3'] = 'Observacion'
        ws['D3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')
        ws['E3'] = 'Organizacion'
        ws['E3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')
        ws['F3'] = 'Fecha de creacion'
        ws['F3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')


        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20

        cont = 4
        for u in claims:
            ws.cell(row = cont, column = 1).value = u.id
            ws.cell(row = cont, column = 2).value = u.by
            ws.cell(row = cont, column = 3).value = u.category
            ws.cell(row = cont, column = 4).value = u.observation
            ws.cell(row = cont, column = 5).value = u.org_name
            ws.cell(row = cont, column = 6).value = u.created
            cont += 1

        excel_name = f'Reporte de reclamos {today}.xlsx'
        response = HttpResponse(content_type = 'application/ms-excel')
        content = 'attachment; filename = {0}'.format(excel_name)
        response['Content-Disposition'] = content
        wb.save(response)
        return response


@login_required
def claim_view(request, pk):
    user_id = request.user.id
    profile_level = Profile.objects.get(user_id = user_id).level

    claim = Claim.objects.get(id=pk)
    
    return render(request,'claims/claim.html', {'claim': claim, 'level': profile_level})


@login_required
def setupclaim_view(request, pk):
    user_id = request.user.id
    profile_level = Profile.objects.get(user_id = user_id).level

    
    org = Org.objects.get(id=pk)

    data = {
       'org': org,
       'level': profile_level,
       }

    if request.method == 'POST':
        claim = Claim()
        claim.category = request.POST.get('category')     
        claim.observation = request.POST.get('observation')     
        claim.state = request.POST.get('state')     
        claim.org = org.id
        claim.org_name = Org.objects.get(id = org.id).name
        claim.by = f'{Profile.objects.get(user_id = user_id).first_name} {Profile.objects.get(user_id = user_id).last_name}'
        claim.save() 

        history_item = Item()
        history_item.action = f'Creacion de reclamo: {claim.org_name}'
        history_item.by = f'{Profile.objects.get(user_id = user_id).first_name} {Profile.objects.get(user_id = user_id).last_name}'
        history_item.save()

        return redirect('claims')

    return render(request,'claims/setupclaim.html', data)


@login_required
def claim_delete_view(request, pk):
    user_id = request.user.id
    claim = Claim.objects.get(id=pk)

    history_item = Item()
    history_item.action = f'Eliminiacion de reclamo: {claim.org_name}'
    history_item.by = f'{Profile.objects.get(user_id = user_id).first_name} {Profile.objects.get(user_id = user_id).last_name}'
    history_item.save()

    claim.delete()

    return redirect('claims')


@login_required
def claim_modify_view(request, pk):
    user_id = request.user.id
    profile_level = Profile.objects.get(user_id = user_id).level
    
    claim = Claim.objects.get(id=pk)

    if request.method == 'POST':
        claim.state = request.POST.get('state')     
        claim.observation = request.POST.get('observation')     
        claim.save() 

        history_item = Item()
        history_item.action = f'Modificacion de reclamo: {claim.org_name}'
        history_item.by = f'{Profile.objects.get(user_id = user_id).first_name} {Profile.objects.get(user_id = user_id).last_name}'
        history_item.save()

        return redirect('claim', claim.id)

    return render(request, 'claims/modify_claim.html', {'claim': claim, 'level': profile_level})