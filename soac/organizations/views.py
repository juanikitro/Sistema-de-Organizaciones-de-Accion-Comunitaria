# Python
from datetime import date, timedelta, datetime

# Django
from django.contrib.auth.decorators import login_required
from django.views.generic import TemplateView
from django.db import IntegrityError
from django.shortcuts import render, redirect

# Models
from organizations.models import Org
from users.models import Profile
from visits.models import Visit
from history.models import Item
from organizations.forms import DocumentForm

# Open Py XL (Para el excel)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side, PatternFill
from django.http.response import HttpResponse

@login_required
def push_soac_view(request):
    
    user_id = request.user.id
    profile_level = Profile.objects.get(user_id = user_id).level

    values = {}
    if request.method == 'POST':
        name = request.POST['name']
        org = Org()
        org.name = name
        org.domain = request.POST['domain']
        org.address = request.POST['address']
        org.dpto = request.POST['dpto']
        org.nhood = request.POST['nhood']
        org.commune = request.POST['commune']
        org.areas = request.POST['areas']
        org.igj = request.POST['igj']
        org.type = request.POST['type']
        org.public = request.POST['public']
        org.postal_code = request.POST['postal_code']
        org.email = request.POST['email']
        org.mobile = request.POST['mobile']
        org.created = date.today()
        org.modified = date.today()
        org.state = 'no-registrada'

        if Org.objects.filter(name=name).first(): 
            return render(request, 'orgs/soac.html', {'error': 'Ya existe una organización con ese nombre, asegurate de no crear la misma', 'values': values, 'level': profile_level})

        org.save()

        history_item = Item()
        history_item.action = f'Creacion de organizacion: {name}'
        history_item.by = f'{Profile.objects.get(user_id = user_id).first_name} {Profile.objects.get(user_id = user_id).last_name}'
        history_item.save()

        return render(request, 'orgs/soac.html', {'alert': 'Organización cargada con exito', 'values': values, 'level': profile_level})

    return render(request, 'orgs/soac.html', {'values': values, 'level': profile_level})

@login_required
def push_roac_view(request):
    user_id = request.user.id
    profile_level = Profile.objects.get(user_id = user_id).level

    values = {}
    if request.method == 'POST':
        name = request.POST['name']
        org = Org()
        org.name = name
        org.domain = request.POST['domain']
        org.address = request.POST['address']
        org.dpto = request.POST['dpto']
        org.nhood = request.POST['nhood']
        org.commune = request.POST['commune']
        org.areas = request.POST['areas']
        org.igj = request.POST['igj']
        org.type = request.POST['type']
        org.public = request.POST['public']
        org.postal_code = request.POST['postal_code']
        org.email = request.POST['email']
        org.mobile = request.POST['mobile']
        org.state = 'pre-activa'

        org.created = date.today()
        org.modified = date.today()
        org.registration_request = date.today()

        if Org.objects.filter(name=name).first(): 
            return render(request, 'orgs/roac.html', {'error': 'Ya existe una organización con ese nombre, asegurate de no crear la misma', 'values': values, 'level': profile_level})

        org.save()

        history_item = Item()
        history_item.action = f'Creacion de organizacion registrada: {name}'
        history_item.by = f'{Profile.objects.get(user_id = user_id).first_name} {Profile.objects.get(user_id = user_id).last_name}'
        history_item.save()

        return redirect('register_roac', org.id)

    return render(request, 'orgs/roac.html', {'values': values, 'level': profile_level})

@login_required
def orgs_view(request):
    user_id = request.user.id
    profile_level = Profile.objects.get(user_id = user_id).level

    global org
    org = Org.objects.all()
    values = {}
    if request.method == 'POST':
        name = request.POST.get('name', False)
        domain = request.POST.get('domain', False)
        address = request.POST.get('address', False)
        nhood = request.POST.get('nhood', False)
        commune = request.POST.get('commune', False)
        areas = request.POST.get('areas', False)
        igj = request.POST.get('igj', False)
        type = request.POST.get('type', False)
        public = request.POST.get('public', False)
        state = request.POST.get('state', False)

        values={
            'name': name,
            'domain': domain,
            'address': address,
            'nhood': nhood,
            'commune': commune,
            'areas': areas,
            'igj': igj,
            'type': type,
            'public': public,
            'state': state,
        }

        org = Org.objects.filter(name__contains=name, domain__contains=domain, address__contains=address, nhood__contains=nhood, commune__contains=commune, areas__contains=areas, igj__contains=igj, type__contains=type, public__contains=public, state__contains=state )

    return render(request, 'orgs/orgs.html', {'org': org, 'values': values, 'level': profile_level})

class Excel_report(TemplateView):
    def get(self, *args, **kwargs):
        today = date.today()

        wb = Workbook()
        ws = wb.active
        ws['A1'] = f'Reporte de organizaciónes del dia: {today}'
        ws['A1'].alignment = Alignment(horizontal = 'center')
        ws['A1'].border = Border(left = Side(border_style = 'thin'), right = Side(border_style = 'thin'), bottom = Side(border_style = 'thin'), top = Side(border_style = 'thin'))
        ws['A1'].font = Font(name = 'Arial', size = 12)
        ws['A1'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')

        ws.merge_cells('A1:N1')

        ws['A3'] = 'id de org.'
        ws['A3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')
        ws['B3'] = 'Nombre'
        ws['B3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')
        ws['C3'] = 'Dominio'
        ws['C3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')
        ws['D3'] = 'Direccion'
        ws['D3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')
        ws['E3'] = 'Departamento'
        ws['E3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')
        ws['F3'] = 'Barrio'
        ws['F3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')
        ws['G3'] = 'Comuna'
        ws['G3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')
        ws['H3'] = 'Areas'
        ws['H3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')
        ws['I3'] = 'IGJ'
        ws['I3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')
        ws['J3'] = 'Tipo'
        ws['J3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')
        ws['K3'] = 'Publico'
        ws['K3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')
        ws['L3'] = 'CP'
        ws['L3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')
        ws['M3'] = 'Contacto'
        ws['M3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')
        ws['N3'] = 'ROAC'
        ws['N3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 20
        ws.column_dimensions['I'].width = 20
        ws.column_dimensions['J'].width = 20
        ws.column_dimensions['K'].width = 20
        ws.column_dimensions['L'].width = 20
        ws.column_dimensions['M'].width = 20

        cont = 4
        for u in org:
            ws.cell(row = cont, column = 1).value = u.id
            ws.cell(row = cont, column = 2).value = u.name
            ws.cell(row = cont, column = 3).value = u.domain
            ws.cell(row = cont, column = 4).value = u.address
            ws.cell(row = cont, column = 5).value = u.dpto
            ws.cell(row = cont, column = 6).value = u.nhood
            ws.cell(row = cont, column = 7).value = u.commune
            ws.cell(row = cont, column = 8).value = u.areas
            ws.cell(row = cont, column = 9).value = u.igj
            ws.cell(row = cont, column = 10).value = u.type
            ws.cell(row = cont, column = 11).value = u.public
            ws.cell(row = cont, column = 12).value = u.postal_code
            ws.cell(row = cont, column = 13).value = u.mobile
            ws.cell(row = cont, column = 14).value = u.roac
            cont += 1

        excel_name = f'Reporte de orgs. {today}.xlsx'
        response = HttpResponse(content_type = 'application/ms-excel')
        content = 'attachment; filename = {0}'.format(excel_name)
        response['Content-Disposition'] = content
        wb.save(response)
        return response

@login_required
def  register_request_view(request, pk):
    user_id = request.user.id
    profile_level = Profile.objects.get(user_id = user_id).level

    selected_org = Org.objects.get(id=pk)

    info = {
        'id': selected_org.id,
        'name': selected_org.name,
    }

    if request.method == 'POST':
        form = DocumentForm(request.POST, request.FILES, instance = selected_org)

        if form.is_valid():
            selected_org.state = 'pre-activa'
            selected_org.msg = ''
            selected_org.registration_request = date.today()
            form.save()

            history_item = Item()
            history_item.action = f'Solicitud de registro: {selected_org.name}'
            history_item.by = f'{Profile.objects.get(user_id = user_id).first_name} {Profile.objects.get(user_id = user_id).last_name}'
            history_item.save()

            return redirect('org', selected_org.id)
            
    else:
        form = DocumentForm()

    return render(request, 'orgs/register_org.html', {'pk': pk, 'info': info, 'form': form, 'level': profile_level})

@login_required
def org_view(request, pk):
    user_id = request.user.id
    profile_level = Profile.objects.get(user_id = user_id).level

    org = Org.objects.get(id=pk)
    visits = Visit.objects.filter(org_name = org.name).order_by('date')

    id = org.id
    name = org.name
    type = org.type
    public = org.public
    areas = org.areas
    address = org.address
    dpto = org.dpto
    postal_code = org.postal_code
    nhood = org.nhood
    commune = org.commune
    domain = org.domain
    email = org.email
    mobile = org.mobile
    state = org.state
    created = org.created
    enrolled = org.enrolled
    renoved = org.renoved
    expiration = org.expiration
    modified = org.modified
    roac = org.roac
    doc = org.doc
    igj = org.igj
    
    context = {
    'level': profile_level,
    'org':org,
    'id':id,
    'name':name,
    'type':type,
    'public':public,
    'areas':areas,
    'domain':domain,
    'commune':commune,
    'nhood':nhood,
    'postal_code':postal_code,
    'dpto':dpto,
    'address':address,
    'email':email,
    'mobile':mobile,
    'state':state,
    'created':created,
    'renoved':renoved,
    'expiration':expiration,
    'modified':modified,
    'roac': roac,
    'doc': doc,
    'enrolled': enrolled,
    'visit': visits,
    'igj': igj,
    'today': datetime.now()

    }

    return render(request, 'orgs/org.html', context)

login_required
def delete_org_view(request, pk):
    user_id = request.user.id

    selected_org = Org.objects.get(id=pk)
    selected_org.delete()

    history_item = Item()
    history_item.action = f'Eliminacion de organizacion: {selected_org.name}'
    history_item.by = f'{Profile.objects.get(user_id = user_id).first_name} {Profile.objects.get(user_id = user_id).last_name}'
    history_item.save()

    return redirect('orgs')

@login_required
def modify_org_view(request, pk):
    user_id = request.user.id
    profile_level = Profile.objects.get(user_id = user_id).level

    selected_org = Org.objects.get(id=pk)
    old_info = {
        'id': selected_org.id,
        'name': selected_org.name,
        'address': selected_org.address,
        'dpto': selected_org.dpto,
        'nhood': selected_org.nhood,
        'commune': selected_org.commune,
        'areas': selected_org.areas,
        'type': selected_org.type,
        'igj': selected_org.igj,
        'email': selected_org.email,
        'mobile': selected_org.mobile,
        'public': selected_org.public,
        'postal_code': selected_org.postal_code,
        'domain': selected_org.domain,
    }

    if request.method == 'POST':

        selected_org.name = request.POST['name']    
        selected_org.address = request.POST['address']    
        selected_org.dpto = request.POST['dpto']    
        selected_org.nhood = request.POST['nhood']    
        selected_org.commune = request.POST['commune']    
        selected_org.areas = request.POST['areas']    
        selected_org.type = request.POST['type']    
        selected_org.igj = request.POST['igj']    
        selected_org.email = request.POST['email']    
        selected_org.mobile = request.POST['mobile']    
        selected_org.public = request.POST['public']    
        selected_org.postal_code = request.POST['postal_code']    
        selected_org.domain = request.POST['domain']    
        selected_org.modified = date.today()

        try:
            selected_org.save()

            history_item = Item()
            history_item.action = f'Modificacion de organizacion: {selected_org.name}'
            history_item.by = f'{Profile.objects.get(user_id = user_id).first_name} {Profile.objects.get(user_id = user_id).last_name}'
            history_item.save()

        except IntegrityError:
            return render(request, 'orgs/modify_org.html', {'old': old_info, 'level': profile_level, 'error': 'Ese nombre ya esta en uso'})
            
        return redirect('orgs')

    return render(request, 'orgs/modify_org.html', {'old': old_info, 'level': profile_level})

@login_required
def down_org_view(request, pk):
    user_id = request.user.id
    selected_org = Org.objects.get(id=pk)

    selected_org.roac = 0
    selected_org.state = 'suspendida'

    try:
        selected_org.save()
        
        history_item = Item()
        history_item.action = f'Suspencion de organizacion: {selected_org.name}'
        history_item.by = f'{Profile.objects.get(user_id = user_id).first_name} {Profile.objects.get(user_id = user_id).last_name}'
        history_item.save()

    except IntegrityError:
        return redirect('org', selected_org.id)

    return redirect('org', selected_org.id)

@login_required
def noregister_org_view(request, pk):
    user_id = request.user.id
    selected_org = Org.objects.get(id=pk)

    selected_org.state = 'no-registrada'
    selected_org.msg= ''

    try:
        selected_org.save()

        history_item = Item()
        history_item.action = f'Devolucion de registro de organizacion: {selected_org.name}'
        history_item.by = f'{Profile.objects.get(user_id = user_id).first_name} {Profile.objects.get(user_id = user_id).last_name}'
        history_item.save()

        return redirect('edit')
    except IntegrityError:
        return redirect('edit')

@login_required
def download_org_view(request, pk):
    selected_org = Org.objects.get(id=pk)

    today = date.today()
    wb = Workbook()
    ws = wb.active
    ws['A1'] = f'Reporte de {selected_org.name} del dia: {today}'
    ws['A1'].alignment = Alignment(horizontal = 'center')
    ws['A1'].border = Border(left = Side(border_style = 'thin'), right = Side(border_style = 'thin'), bottom = Side(border_style = 'thin'), top = Side(border_style = 'thin'))
    ws['A1'].font = Font(name = 'Arial', size = 12)
    ws['A1'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')

    ws.merge_cells('A1:N1')

    ws['A3'] = 'id de org.'
    ws['A3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')
    ws['B3'] = 'Nombre'
    ws['B3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')
    ws['C3'] = 'Dominio'
    ws['C3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')
    ws['D3'] = 'Direccion'
    ws['D3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')
    ws['E3'] = 'Departamento'
    ws['E3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')
    ws['F3'] = 'Barrio'
    ws['F3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')
    ws['G3'] = 'Comuna'
    ws['G3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')
    ws['H3'] = 'Areas'
    ws['H3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')
    ws['I3'] = 'IGJ'
    ws['I3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')
    ws['J3'] = 'Tipo'
    ws['J3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')
    ws['K3'] = 'Publico'
    ws['K3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')
    ws['L3'] = 'CP'
    ws['L3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')
    ws['M3'] = 'Contacto'
    ws['M3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')
    ws['N3'] = 'ROAC'
    ws['N3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 20
    ws.column_dimensions['K'].width = 20
    ws.column_dimensions['L'].width = 20
    ws.column_dimensions['M'].width = 20

    cont = 4
    if selected_org:
        ws.cell(row = cont, column = 1).value = selected_org.id
        ws.cell(row = cont, column = 2).value = selected_org.name
        ws.cell(row = cont, column = 3).value = selected_org.domain
        ws.cell(row = cont, column = 4).value = selected_org.address
        ws.cell(row = cont, column = 5).value = selected_org.dpto
        ws.cell(row = cont, column = 6).value = selected_org.nhood
        ws.cell(row = cont, column = 7).value = selected_org.commune
        ws.cell(row = cont, column = 8).value = selected_org.areas
        ws.cell(row = cont, column = 9).value = selected_org.igj
        ws.cell(row = cont, column = 10).value = selected_org.type
        ws.cell(row = cont, column = 11).value = selected_org.public
        ws.cell(row = cont, column = 12).value = selected_org.postal_code
        ws.cell(row = cont, column = 13).value = selected_org.mobile
        ws.cell(row = cont, column = 14).value = selected_org.roac
        cont += 1

    excel_name = f'Reporte de {selected_org.name} {today}.xlsx'
    response = HttpResponse(content_type = 'application/ms-excel')
    content = 'attachment; filename = {0}'.format(excel_name)
    response['Content-Disposition'] = content
    wb.save(response)
    return response

@login_required
def signing_org_view(request, pk):
    user_id = request.user.id
    profile_level = Profile.objects.get(user_id = user_id).level

    selected_org = Org.objects.get(id=pk)

    info = {
        'id': selected_org.id,
        'name': selected_org.name,
        'doc': selected_org.doc
    }

    if request.method == 'POST':
        form = DocumentForm(request.POST, request.FILES, instance = selected_org)

        if form.is_valid():
            selected_org.state = 'activa'
            selected_org.roac = 'Si'
            selected_org.expiration = date.today() + timedelta(days=730)
            selected_org.msg = ''
            form.save()

            history_item = Item()
            history_item.action = f'Activacion de organizacion: {selected_org.name}'
            history_item.by = f'{Profile.objects.get(user_id = user_id).first_name} {Profile.objects.get(user_id = user_id).last_name}'
            history_item.save()

            return redirect('org', selected_org.id)
            
    else:
        form = DocumentForm()

    return render(request, 'orgs/register_org.html', {'pk': pk, 'info': info, 'form': form})