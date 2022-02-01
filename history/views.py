from datetime import date

#Django 
from django.views.generic import TemplateView
from django.shortcuts import render
from django.contrib.auth.decorators import login_required

#Models
from users.models import Profile
from history.models import Item

# Open Py XL (Para el excel)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side, PatternFill
from django.http.response import HttpResponse

@login_required
def history_view(request):
    user_id = request.user.id
    profile_level = Profile.objects.get(user_id = user_id).level

    global items
    values = {}
    items = Item.objects.all().order_by('-date')

    if request.method == 'POST':
        action = request.POST.get('action', False)
        user = request.POST.get('user', False)
        date = request.POST.get('date', False)

        values={
            'action': action,
            'user': user,
            'date': date,
        }

        items = Item.objects.filter(action__contains=action, date__contains=date, by__contains=user).order_by('-date')


    return render(request, 'history/history.html', {'level': profile_level, 'items': items, 'values':values})

class Excel_report(TemplateView):
    def get(self, *args, **kwargs):
        today = date.today()

        wb = Workbook()
        ws = wb.active
        ws['A1'] = f'Reporte de historial del dia: {today}'
        ws['A1'].alignment = Alignment(horizontal = 'center')
        ws['A1'].border = Border(left = Side(border_style = 'thin'), right = Side(border_style = 'thin'), bottom = Side(border_style = 'thin'), top = Side(border_style = 'thin'))
        ws['A1'].font = Font(name = 'Arial', size = 12)
        ws['A1'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')

        ws.merge_cells('A1:D1')

        ws['A3'] = 'id de item'
        ws['A3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')
        ws['B3'] = 'Accion'
        ws['B3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')
        ws['C3'] = 'Usuario'
        ws['C3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')
        ws['D3'] = 'Fecha'
        ws['D3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')


        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20

        cont = 4
        for u in items:
            ws.cell(row = cont, column = 1).value = u.id
            ws.cell(row = cont, column = 2).value = u.action
            ws.cell(row = cont, column = 3).value = u.by
            ws.cell(row = cont, column = 4).value = u.date
            cont += 1

        excel_name = f'Reporte de historial {today}.xlsx'
        response = HttpResponse(content_type = 'application/ms-excel')
        content = 'attachment; filename = {0}'.format(excel_name)
        response['Content-Disposition'] = content
        wb.save(response)
        return response