#Django & python
from datetime import datetime, date
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.core.mail import send_mail
from django.conf import settings
from django.views.generic import TemplateView

# Open Py XL (Para el excel)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side, PatternFill
from django.http.response import HttpResponse

#Modelos
from activities.models import Activity
from organizations.models import Org
from history.models import Item
from users.models import Profile

@login_required
def activities_view(request):
    ''' Actividades
    Calendario de actividades
    Creacion de actividades
    Listado de actividadeas
    nashe '''
    user_id = request.user.id
    profile_level = Profile.objects.get(user_id = user_id).level

    activities = Activity.objects.all().order_by('date')
    today = datetime.now().date()
    year = today.year
    orgs = Org.objects.all()

    if request.method == 'POST':
        activity = Activity()
        activity.activity_type = request.POST.get('activity_type')     
        activity.date = request.POST.get('date')     
        activity.hour = request.POST.get('hour')
        if request.POST.get('allday') == 'on':
            activity.allday = 'Si'
        else: 
            activity.allday = 'No'    
        activity.save() 

        orgs_id = request.POST.getlist('orgs')
        emails = []
        communes = []
        orgs_names = []

        history_item = Item()
        history_item.action = f'Creacion de actividad: {activity.activity_type}'
        history_item.by = f'{Profile.objects.get(user_id = user_id).first_name} {Profile.objects.get(user_id = user_id).last_name}'
        history_item.save()

        for u in orgs_id:
            org = Org.objects.get(id=u)
            activity.orgs.add(org)
            emails.append(org.email)
            communes.append(org.commune)
            orgs_names.append(org.name)
        
        activity.orgs_names = orgs_names    
        activity.communes = communes    
        activity.save() 

        if request.POST.get('notify') == 'on':
            subject = f'SOAC: Actividad: {activity.activity_type}'
            link = f'http://172.31.67.157/activities/{activity.id}/' #FIXME: Cambiar cuando existan los servers
            email_from = settings.EMAIL_HOST_USER
            message = f'''Hola! Te contacto desde SOAC porque {request.user.first_name} {request.user.last_name} creo la actividad: {activity.activity_type}. 
            Fecha: {activity.date}
            Hora: {activity.hour}
            Tipo: {activity.activity_type}
            Podes ver mas sobre esta entrando al siguiente link:
            {link}
            
            DG Relaciones con la Comunidad
            SS de Gestión Comunal
            '''

            send_mail(subject, message, email_from, emails)

        return redirect('activities')

    return render(request,'activities/activities.html', {'today': datetime.now(), 'activity': activities, 'orgs': orgs, 'year': year, 'level': profile_level})


@login_required
def activity_view(request, pk):
    ''' Perfil de actividad '''
    user_id = request.user.id
    profile_level = Profile.objects.get(user_id = user_id).level
    
    activity = Activity.objects.get(id=pk)

    orgs = ''
    i = activity.orgs.all()
    ids = i.values_list('pk', flat=True)
    for u in ids:
        org = Org.objects.get(id=u)
        orgs_names = (f'{org.name}, ')
        orgs = orgs + orgs_names
    
    return render(request,'activities/activity.html', {'activity':activity, 'orgs':orgs, 'level': profile_level})


@login_required
def activity_delete_view(request, pk):
    ''' Eliminar actividad '''
    user_id = request.user.id
    activity = Activity.objects.get(id=pk)

    emails = []
    subject = f'SOAC: Eliminacion de actividad "{activity.activity_type}"'
    email_from = settings.EMAIL_HOST_USER
    message = f'''Hola! Te contacto desde SOAC para informarte que la actividad "{activity.activity_type}" ha sido eliminado.
    
    DG Relaciones con la Comunidad
    SS de Gestión Comunal
    '''

    i = activity.orgs.all()
    ids = i.values_list('pk', flat=True)
    for u in ids:
        org = Org.objects.get(id=u)
        emails.append(org.email)

    send_mail(subject, message, email_from, emails)

    history_item = Item()
    history_item.action = f'Eliminacion de actividad: {activity.activity_type}'
    history_item.by = f'{Profile.objects.get(user_id = user_id).first_name} {Profile.objects.get(user_id = user_id).last_name}'
    history_item.save()

    activity.delete()
    
    return redirect('activities')


@login_required
def activity_modify_view(request, pk):
    ''' Modificar actividad '''
    user_id = request.user.id
    profile_level = Profile.objects.get(user_id = user_id).level

    activity = Activity.objects.get(id=pk)

    if request.method == 'POST':
        activity.activity_type = request.POST.get('activity_type')     
        activity.date = request.POST.get('date')     
        activity.hour = request.POST.get('hour')     
        activity.save()

        emails = []
        subject = f'SOAC: Modificacion al actividad "{activity.activity_type}"'
        link = f'http://172.31.67.157/activities/{activity.id}/' #FIXME: Cambiar cuando existan los servers
        email_from = settings.EMAIL_HOST_USER
        message = f'''Hola! Te contacto desde SOAC porque se ha actualizado la actividad "{request.POST.get('activity_type')}". 
        Fecha: {request.POST.get('date')}
        Hora: {request.POST.get('hour')}
        Podes ver mas sobre esta entrando al siguiente link:
        {link}
        
        DG Relaciones con la Comunidad
        SS de Gestión Comunal
        '''

        i = activity.orgs.all()
        ids = i.values_list('pk', flat=True)
        for u in ids:
            org = Org.objects.get(id=u)
            emails.append(org.email)

        send_mail(subject, message, email_from, emails)

        history_item = Item()
        history_item.action = f'Modificacion de actividad: {activity.activity_type}'
        history_item.by = f'{Profile.objects.get(user_id = user_id).first_name} {Profile.objects.get(user_id = user_id).last_name}'
        history_item.save()

        activity.save()
        return redirect('activity', activity.id)

    return render(request, 'activities/modify_activ.html', {'activity': activity, 'level': profile_level})


@login_required
def activitiesreport_view(request):
    user_id = request.user.id
    profile_level = Profile.objects.get(user_id = user_id).level

    global activities
    activities = Activity.objects.all()

    values = {}
    if request.method == 'POST':
        activity_type = request.POST.get('activity_type', False)
        date = request.POST.get('date', False)
        hour = request.POST.get('hour', False)
        communes = request.POST.get('communes', False)
        orgs_names = request.POST.get('orgs_names', False)

        values={
            'date': date,
            'hour': hour,
            'activity_type': activity_type,
        }

        activities = Activity.objects.filter(orgs_names__contains = orgs_names, communes__contains = communes, date__contains = date, hour__contains = hour, activity_type__contains = activity_type)

    return render(request, 'activities/activitiesreport.html', {'activities': activities, 'values': values, 'level': profile_level})


class Activities_excel_report(TemplateView):
    '''Export de usuarios
    Se crea el excel con los datos de los usuarios
    Se utilizan los usuarios filtrados por la variable global profile '''

    def get(self, *args, **kwargs):
        today = date.today()

        wb = Workbook()
        ws = wb.active
        ws['A1'] = f'Reporte de actividades del dia: {today}'
        ws['A1'].alignment = Alignment(horizontal = 'center')
        ws['A1'].border = Border(left = Side(border_style = 'thin'), right = Side(border_style = 'thin'), bottom = Side(border_style = 'thin'), top = Side(border_style = 'thin'))
        ws['A1'].font = Font(name = 'Arial', size = 12)
        ws['A1'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')

        ws.merge_cells('A1:E1')

        ws['A3'] = 'ID de actividad'
        ws['A3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')
        ws['B3'] = 'Fecha'
        ws['B3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')
        ws['C3'] = 'Hora'
        ws['C3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')
        ws['D3'] = 'Todo el dia'
        ws['D3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')
        ws['E3'] = 'Tipo'
        ws['E3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20

        cont = 4
        for u in activities:
            ws.cell(row = cont, column = 1).value = u.id
            ws.cell(row = cont, column = 2).value = u.date
            ws.cell(row = cont, column = 3).value = u.hour
            ws.cell(row = cont, column = 4).value = u.allday
            ws.cell(row = cont, column = 5).value = u.activity_type
            cont += 1

        excel_name = f'Reporte de actividades {today}.xlsx'
        response = HttpResponse(content_type = 'application/ms-excel')
        content = 'attachment; filename = {0}'.format(excel_name)
        response['Content-Disposition'] = content
        wb.save(response)
        return response