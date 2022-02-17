#Django & python
from ast import If
from datetime import datetime, date
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.core.mail import send_mail
from django.views.generic import TemplateView
from django.conf import settings

#Modelos
from visits.models import Visit, Act
from claims.models import Claim
from organizations.models import Org
from history.models import Item
from users.models import Profile

# Open Py XL (Para el excel)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side, PatternFill
from django.http.response import HttpResponse

# Visitas
@login_required
def visits_view(request):
    ''' Visitas
    Calendario de visitas
    Creacion de visitas
    Listado de visitas '''

    user_id = request.user.id
    profile_level = Profile.objects.get(user_id = user_id).level
    visits = Visit.objects.all().order_by('date')
    orgs = Org.objects.all()
    today = datetime.now().date()
    year = today.year
   
    data = {
       'visit': visits,
       'orgs': orgs,
       'year' : year,
       'level': profile_level,
       'today': datetime.now()
       }

    if request.method == 'POST':
        visit = Visit()
        visit.date = request.POST.get('date')     
        visit.hour = request.POST.get('hour')     
        visit.observation = request.POST.get('observation')     
        visit.org = request.POST.get('org')  
        visit.org_name = Org.objects.get(id = request.POST.get('org')).name
        if request.POST.get('allday') == 'on':
            visit.allday = 'Si'
        else: 
            visit.allday = 'No'
        email = Org.objects.get(id = request.POST.get('org')).email
        visit.save() 
    
        emails  = [email]
        subject = f'SOAC: Visita del dia {visit.date}'
        link = f'http://172.31.67.157/visits/{visit.id}/' #FIXME: Cambiar cuando existan los servers
        email_from = settings.EMAIL_HOST_USER
        message = f'''Hola! Te contacto desde SOAC porque {request.user.first_name} {request.user.last_name} creo una visita. 
        Fecha: {visit.date}
        Hora: {visit.hour}
        Observacion: {visit.observation}
        Podes ver mas sobre esta entrando al siguiente link:
        {link}
        
        DG Relaciones con la Comunidad
        SS de Gestión Comunal
        '''

        send_mail(subject, message, email_from, emails)

        history_item = Item()
        history_item.action = f'Creacion de visita: {visit.org_name}'
        history_item.by = f'{Profile.objects.get(user_id = user_id).first_name} {Profile.objects.get(user_id = user_id).last_name}'
        history_item.save()

        return redirect('visits')

    return render(request,'visits/visits.html', data)


@login_required
def visit_view(request, pk):
    ''' Perfil de visita '''
    user_id = request.user.id
    profile_level = Profile.objects.get(user_id = user_id).level

    visit = Visit.objects.get(id=pk)
    
    return render(request,'visits/visit.html', {'visit':visit,'today': datetime.now(), 'level': profile_level})


@login_required
def visit_delete_view(request, pk):
    ''' Eliminar visita '''
    user_id = request.user.id
    visit = Visit.objects.get(id=pk)
    email = Org.objects.get(id = visit.org).email

    emails = [email]
    subject = f'SOAC: Eliminacion de visita del dia {visit.date}'
    email_from = settings.EMAIL_HOST_USER
    message = f'''Hola! Te contacto desde SOAC para informarte que la visita del dia: {visit.date} ha sido eliminada.
    
    DG Relaciones con la Comunidad
    SS de Gestión Comunal
    '''

    send_mail(subject, message, email_from, emails)

    history_item = Item()
    history_item.action = f'Eliminiacion de visita: {visit.org_name}'
    history_item.by = f'{Profile.objects.get(user_id = user_id).first_name} {Profile.objects.get(user_id = user_id).last_name}'
    history_item.save()

    visit.delete()

    return redirect('visits')


@login_required
def visit_modify_view(request, pk):
    ''' Modifivar visita '''
    user_id = request.user.id
    profile_level = Profile.objects.get(user_id = user_id).level
    
    visit = Visit.objects.get(id=pk)

    if request.method == 'POST':
        visit.date = request.POST.get('date')     
        visit.hour = request.POST.get('hour')     
        visit.observation = request.POST.get('observation')     
        email = Org.objects.get(id = visit.org).email
        visit.save() 
    
        emails  = [email]
        subject = f'SOAC: Modificacion a la visita del dia {visit.date}'
        link = f'http://172.31.67.157/visits/{visit.id}/' #FIXME: Cambiar cuando existan los servers
        email_from = settings.EMAIL_HOST_USER
        message = f'''Hola! Te contacto desde SOAC porque se ha actualizado la visita del dia {request.POST.get('date')}. 
        Fecha: {request.POST.get('date')}
        Hora: {request.POST.get('hour')}
        Observacion: {request.POST.get('observation')}
        Podes ver mas sobre este entrando al siguiente link:
        {link}
        
        DG Relaciones con la Comunidad
        SS de Gestión Comunal
        '''

        send_mail(subject, message, email_from, emails)

        history_item = Item()
        history_item.action = f'Modificacion de visita: {visit.org_name}'
        history_item.by = f'{Profile.objects.get(user_id = user_id).first_name} {Profile.objects.get(user_id = user_id).last_name}'
        history_item.save()

        return redirect('visit', visit.id)

    return render(request, 'visits/modify_visit.html', {'visit': visit, 'level': profile_level})


#Actas de visitas
@login_required
def create_act_view(request, pk):
    ''' Crear acta de visitas dentro de la visita '''
    user_id = request.user.id
    profile_level = Profile.objects.get(user_id = user_id).level

    visit = Visit.objects.get(id = pk)
   
    data = {
       'level': profile_level,
       'visit': visit
       }

    if request.method == 'POST':
        act = Act()
        act.date = visit.date
        act.agent = request.POST.get('agent')     
        act.receptor_name = request.POST.get('receptor_name')
        act.receptor_charge = request.POST.get('receptor_charge')
        act.beneficiaries = request.POST.get('beneficiaries')
        act.partners = request.POST.get('partners')
        act.tasks = request.POST.get('tasks')
        act.subsidies = request.POST.get('subsidies')
        act.subsidies_what = request.POST.get('subsidies_what')
        act.links = request.POST.get('links')
        act.visit = visit   
        act.tasks = request.POST.getlist('tasks')

        print(act.tasks)

        # task_list = ''
        # tasks = request.POST.getlist('tasks')
        # for u in tasks:
        #     task_list.add(org)
        # act.tasks = task_list

        act.claim_exist = 'No'
        if request.POST.get('category') != '':
            claim = Claim()
            claim.category = request.POST.get('category')     
            claim.observation = request.POST.get('observation')     
            claim.state = request.POST.get('state')     
            claim.org = visit.org
            claim.org_name = Org.objects.get(id = visit.org).name
            claim.by = f'{Profile.objects.get(user_id = user_id).first_name} {Profile.objects.get(user_id = user_id).last_name}'
            claim.save() 
            act.claim = claim
            act.claim_exist = 'Si'

        act.save() 

        visit.act_id = act.id
        visit.save()

        history_item = Item()
        history_item.action = f'Creacion de acta a visita: {visit.id}'
        history_item.by = f'{Profile.objects.get(user_id = user_id).first_name} {Profile.objects.get(user_id = user_id).last_name}'
        history_item.save()

        return redirect('visit', pk)

    return render(request,'visits/create_act.html', data)


@login_required
def act_view(request, pk):
    ''' Perfil de acta de visita '''
    user_id = request.user.id
    profile_level = Profile.objects.get(user_id = user_id).level

    visit = Visit.objects.get(id = pk)
    act = Act.objects.get(visit_id = pk)
    claim = ''
    if act.claim_exist == 'Si':
        claim = Claim.objects.get(id = act.claim_id)
   
    data = {
       'level': profile_level,
       'visit': visit,
       'act': act,
       'claim': claim
       }

    return render(request,'visits/act.html', data)


@login_required
def visitsreport_view(request):
    user_id = request.user.id
    profile_level = Profile.objects.get(user_id = user_id).level

    global visits
    visits = Visit.objects.all()
    orgs = Org.objects.all()

    values = {}
    if request.method == 'POST':
        date = request.POST.get('date', False)
        hour = request.POST.get('hour', False)
        observation = request.POST.get('observation', False)
        org = request.POST.get('org', False)

        values={
            'date': date,
            'hour': hour,
            'org': org,
            'observation': observation,
        }

        visits = Visit.objects.filter(date__contains = date, hour__contains = hour, observation__contains = observation, org__contains = org,)

    return render(request, 'visits/visitsreport.html', {'visits': visits, 'orgs': orgs, 'values': values, 'level': profile_level})


class Visits_excel_report(TemplateView):
    '''Export de usuarios
    Se crea el excel con los datos de los usuarios
    Se utilizan los usuarios filtrados por la variable global profile '''

    def get(self, *args, **kwargs):
        today = date.today()

        wb = Workbook()
        ws = wb.active
        ws['A1'] = f'Reporte de visitas del dia: {today}'
        ws['A1'].alignment = Alignment(horizontal = 'center')
        ws['A1'].border = Border(left = Side(border_style = 'thin'), right = Side(border_style = 'thin'), bottom = Side(border_style = 'thin'), top = Side(border_style = 'thin'))
        ws['A1'].font = Font(name = 'Arial', size = 12)
        ws['A1'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')

        ws.merge_cells('A1:G1')

        ws['A3'] = 'ID de visita'
        ws['A3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')
        ws['B3'] = 'Fecha'
        ws['B3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')
        ws['C3'] = 'Hora'
        ws['C3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')
        ws['D3'] = 'Todo el dia'
        ws['D3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')
        ws['E3'] = 'Observacion'
        ws['E3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')
        ws['F3'] = 'ID de organizacion'
        ws['F3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')
        ws['G3'] = 'Nombre de organizacion'
        ws['G3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 20

        cont = 4
        for u in visits:
            ws.cell(row = cont, column = 1).value = u.id
            ws.cell(row = cont, column = 2).value = u.date
            ws.cell(row = cont, column = 3).value = u.hour
            ws.cell(row = cont, column = 4).value = u.allday
            ws.cell(row = cont, column = 5).value = u.observation
            ws.cell(row = cont, column = 6).value = u.org
            ws.cell(row = cont, column = 7).value = u.org_name
            cont += 1

        excel_name = f'Reporte de visitas {today}.xlsx'
        response = HttpResponse(content_type = 'application/ms-excel')
        content = 'attachment; filename = {0}'.format(excel_name)
        response['Content-Disposition'] = content
        wb.save(response)
        return response