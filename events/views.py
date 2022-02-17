#Django & python
from datetime import datetime,date
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.core.mail import send_mail
from django.conf import settings
from django.views.generic import TemplateView

# Open Py XL (Para el excel)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side, PatternFill
from django.http.response import HttpResponse

#Modelos
from .models import Event
from activities.models import Activity
from visits.models import Visit
from history.models import Item
from organizations.models import Org
from users.models import Profile

@login_required
def general_calendar_view(request):
    ''' Calendario de eventos / actividades / eventos '''

    user_id = request.user.id
    profile_level = Profile.objects.get(user_id = user_id).level

    events = Event.objects.all()
    activities = Activity.objects.all()
    visits = Visit.objects.all()
    today = datetime.now().date()
    year = today.year
   
    data = {
       'event': events,
       'activity': activities,
       'visit': visits,
       'year' : year,
       'level': profile_level
       }

    return render(request,'events/calendar.html', data)


@login_required
def events_view(request):
    ''' Eventos 
    Calendario de eventos
    Creacion de eventos
    Listado de eventos '''
    user_id = request.user.id
    profile_level = Profile.objects.get(user_id = user_id).level

    events = Event.objects.all().order_by('date')
    orgs = Org.objects.all()
    today = datetime.now().date()
    year = today.year

    data = {
       'event': events,
       'orgs': orgs,
       'year' : year,
       'level': profile_level,
       'today': datetime.now()
       }

    if request.method == 'POST':
        event = Event()
        event.event_name = request.POST.get('event_name')     
        event.date = request.POST.get('date')     
        event.hour = request.POST.get('hour')     
        event.spot = request.POST.get('spot')
        if request.POST.get('allday') == 'on':
            event.allday = 'Si'
        else: 
            event.allday = 'No'
        event.save() 

        history_item = Item()
        history_item.action = f'Creacion de evento: {event.event_name}'
        history_item.by = f'{Profile.objects.get(user_id = user_id).first_name} {Profile.objects.get(user_id = user_id).last_name}'
        history_item.save()

        emails = []
        communes = []
        orgs_names = []

        orgs_id = request.POST.getlist('orgs')
        for u in orgs_id:
            org = Org.objects.get(id=u)
            event.orgs.add(org)
            emails.append(org.email)
            orgs_names.append(org.name)
            communes.append(org.commune)

        event.orgs_names = orgs_names
        event.communes = communes
        event.save() 

        if request.POST.get('notify') == 'on':
            subject = f'SOAC: Invitacion al evento "{event.event_name}"'
            link = f'http://172.31.67.157/events/{event.id}/' #FIXME: Cambiar cuando existan los servers
            email_from = settings.EMAIL_HOST_USER
            message = f'''Hola! Te contacto desde SOAC porque {request.user.first_name} {request.user.last_name} te invito al evento "{event.event_name}". 
            Fecha: {event.date}
            Hora: {event.hour}
            Lugar: {event.spot}
            Podes ver mas sobre este entrando al siguiente link:
            {link}
            
            DG Relaciones con la Comunidad
            SS de Gestión Comunal
            '''

            send_mail(subject, message, email_from, emails)

        return redirect('events')

    return render(request,'events/events.html', data)


@login_required
def event_view(request, pk):
    ''' Perfil de evento '''
    user_id = request.user.id
    profile_level = Profile.objects.get(user_id = user_id).level

    event = Event.objects.get(id=pk)

    orgs = ''
    i = event.orgs.all()
    ids = i.values_list('pk', flat=True)
    for u in ids:
        org = Org.objects.get(id=u)
        orgs_names = (f'{org.name}, ')
        orgs = orgs + orgs_names
    
    return render(request,'events/event.html', {'event':event, 'orgs':orgs, 'level': profile_level})


@login_required
def event_delete_view(request, pk):
    ''' Eliminar evento '''
    user_id = request.user.id
    event = Event.objects.get(id=pk)

    emails = []
    subject = f'SOAC: Eliminacion de evento "{event.event_name}"'
    email_from = settings.EMAIL_HOST_USER
    message = f'''Hola! Te contacto desde SOAC para informarte que el evento "{event.event_name}" ha sido eliminado.
    
    DG Relaciones con la Comunidad
    SS de Gestión Comunal
    '''

    i = event.orgs.all()
    ids = i.values_list('pk', flat=True)
    for u in ids:
        org = Org.objects.get(id=u)
        emails.append(org.email)

    send_mail(subject, message, email_from, emails)

    history_item = Item()
    history_item.action = f'Eliminacion de evento: {event.event_name}'
    history_item.by = f'{Profile.objects.get(user_id = user_id).first_name} {Profile.objects.get(user_id = user_id).last_name}'
    history_item.save()

    event.delete()
    
    return redirect('events')


@login_required
def event_modify_view(request, pk):
    ''' Modificar evento '''
    user_id = request.user.id
    profile_level = Profile.objects.get(user_id = user_id).level
    
    event = Event.objects.get(id=pk)

    if request.method == 'POST':
        event.event_name = request.POST.get('event_name')     
        event.date = request.POST.get('date')     
        event.hour = request.POST.get('hour')     
        event.spot = request.POST.get('spot')     
        event.save()

        history_item = Item()
        history_item.action = f'Modificacion de actividad: {event.event_name}'
        history_item.by = f'{Profile.objects.get(user_id = user_id).first_name} {Profile.objects.get(user_id = user_id).last_name}'
        history_item.save()

        emails = []
        subject = f'SOAC: Modificacion al evento "{event.event_name}"'
        link = f'http://172.31.67.157/events/{event.id}/' #FIXME: Cambiar cuando existan los servers
        email_from = settings.EMAIL_HOST_USER
        message = f'''Hola! Te contacto desde SOAC porque se ha actualizado el evento "{request.POST.get('event_name')}". 
        Fecha: {request.POST.get('date')}
        Hora: {request.POST.get('hour')}
        Lugar: {request.POST.get('spot')}
        Podes ver mas sobre este entrando al siguiente link:
        {link}
        
        DG Relaciones con la Comunidad
        SS de Gestión Comunal
        '''

        i = event.orgs.all()
        ids = i.values_list('pk', flat=True)
        for u in ids:
            org = Org.objects.get(id=u)
            emails.append(org.email)

        send_mail(subject, message, email_from, emails)

        event.save()
        return redirect('event', event.id)

    return render(request, 'events/modify_event.html', {'event': event, 'level': profile_level})


@login_required
def eventsreport_view(request):
    user_id = request.user.id
    profile_level = Profile.objects.get(user_id = user_id).level

    global events
    events = Event.objects.all()

    values = {}
    if request.method == 'POST':
        event_name = request.POST.get('event_name', False)
        date = request.POST.get('date', False)
        hour = request.POST.get('hour', False)
        spot = request.POST.get('spot', False)
        communes = request.POST.get('communes', False)
        orgs_names = request.POST.get('orgs_names', False)

        values={
            'date': date,
            'hour': hour,
            'event_name': event_name,
            'spot': spot,
            'communes': communes,
            'orgs_names': orgs_names,
        }

        print(event_name, date, hour, spot, communes, orgs_names)
        events = Event.objects.filter(date__contains = date,orgs_names__contains = orgs_names, hour__contains = hour, event_name__contains = event_name, spot__contains = spot, communes__contains = communes)

    return render(request, 'events/eventsreport.html', {'events': events, 'values': values, 'level': profile_level})


class Events_excel_report(TemplateView):
    '''Export de usuarios
    Se crea el excel con los datos de los usuarios
    Se utilizan los usuarios filtrados por la variable global profile '''

    def get(self, *args, **kwargs):
        today = date.today()

        wb = Workbook()
        ws = wb.active
        ws['A1'] = f'Reporte de eventos del dia: {today}'
        ws['A1'].alignment = Alignment(horizontal = 'center')
        ws['A1'].border = Border(left = Side(border_style = 'thin'), right = Side(border_style = 'thin'), bottom = Side(border_style = 'thin'), top = Side(border_style = 'thin'))
        ws['A1'].font = Font(name = 'Arial', size = 12)
        ws['A1'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')

        ws.merge_cells('A1:E1')

        ws['A3'] = 'ID de evento'
        ws['A3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')
        ws['B3'] = 'Fecha'
        ws['B3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')
        ws['C3'] = 'Hora'
        ws['C3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')
        ws['D3'] = 'Todo el dia'
        ws['D3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')
        ws['E3'] = 'Nombre'
        ws['E3'].fill = PatternFill(start_color = 'ffc107', end_color = 'f3b600', fill_type='solid')

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20

        cont = 4
        for u in events:
            ws.cell(row = cont, column = 1).value = u.id
            ws.cell(row = cont, column = 2).value = u.date
            ws.cell(row = cont, column = 3).value = u.hour
            ws.cell(row = cont, column = 4).value = u.allday
            ws.cell(row = cont, column = 5).value = u.event_name
            cont += 1

        excel_name = f'Reporte de eventos {today}.xlsx'
        response = HttpResponse(content_type = 'application/ms-excel')
        content = 'attachment; filename = {0}'.format(excel_name)
        response['Content-Disposition'] = content
        wb.save(response)
        return response